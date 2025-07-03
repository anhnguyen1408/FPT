import os
import time
import shutil
import xml.etree.ElementTree as ET
from urllib.parse import urlparse

import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook, Workbook


def setup_browser(download_dir):
    os.makedirs(download_dir, exist_ok=True)
    options = Options()
    prefs = {
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "download.default_directory": download_dir,
        "plugins.always_open_pdf_externally": True,
        "profile.default_content_settings.popups": 0,
        "safebrowsing.enabled": True,
        "profile.default_content_setting_values.automatic_downloads": 1
    }
    options.add_experimental_option("prefs", prefs)
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-popup-blocking")
    service = Service()  # Add executable_path if needed
    driver = webdriver.Chrome(service=service, options=options)
    wait = WebDriverWait(driver, 15)
    return driver, wait


def lookup_invoice(driver, wait, tax_code, lookup_code, url):
    try:
        driver.get(url)
        if "tracuuhoadon.fpt.com.vn" in url:
            tax_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='MST bên bán']")))
            tax_input.clear()
            tax_input.send_keys(tax_code.strip())

            code_input = driver.find_element(By.XPATH, "//input[@placeholder='Mã tra cứu hóa đơn']")
            code_input.clear()
            code_input.send_keys(lookup_code.strip())

            search_btn = driver.find_element(By.XPATH, "//button[contains(@class, 'webix_button') and contains(text(), 'Tra cứu')]")
            search_btn.click()

        elif "meinvoice.vn/tra-cuu/" in url:
            code_input = wait.until(EC.presence_of_element_located((By.NAME, "txtCode")))
            code_input.clear()
            code_input.send_keys(lookup_code.strip())

            search_btn = wait.until(EC.element_to_be_clickable((By.ID, "btnSearchInvoice")))
            search_btn.click()

        elif "van.ehoadon.vn" in url:
            invoice_input = wait.until(EC.presence_of_element_located((By.ID, "txtInvoiceCode")))
            invoice_input.clear()
            invoice_input.send_keys(lookup_code.strip())

            search_btn = driver.find_element(By.CLASS_NAME, "btnSearch")
            search_btn.click()

        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))

    except TimeoutException:
        print(f"Lookup failed for: {lookup_code}")


def download_xml(driver, wait, base_dir, url, lookup_code):
    try:
        if "tracuuhoadon.fpt.com.vn" in url:
            btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[span[contains(@class, 'mdi-xml')] and contains(text(), 'Tải XML')]")))
            btn.click()

        elif "meinvoice.vn/tra-cuu/" in url:
            btn = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "download")))
            driver.execute_script("arguments[0].click();", btn)
            xml_btn = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "txt-download-xml")))
            xml_btn.click()

        elif "van.ehoadon.vn" in url:
            wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "frameViewInvoice")))
            btn = wait.until(EC.presence_of_element_located((By.ID, "btnDownload")))
            ActionChains(driver).move_to_element(btn).perform()
            driver.execute_script("document.querySelector('#divDownloads .dropdown-menu').style.display='block';")
            xml_btn = wait.until(EC.element_to_be_clickable((By.ID, "LinkDownXML")))
            xml_btn.click()
            driver.switch_to.default_content()

    except TimeoutException:
        print(f"Download failed for: {lookup_code}")
        return None

    domain = urlparse(url).netloc.replace("www.", "").replace(":", "_")
    domain_folder = os.path.join(base_dir, domain)
    os.makedirs(domain_folder, exist_ok=True)

    for _ in range(10):
        files = os.listdir(base_dir)
        for file in files:
            if file.endswith(".xml") and not file.endswith(".crdownload"):
                src = os.path.join(base_dir, file)
                dest = os.path.join(domain_folder, f"{lookup_code}.xml")
                shutil.move(src, dest)
                return dest
        time.sleep(1)
    return None


def parse_xml(xml_path):
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
        invoice_node = None

        hdon_node = root.find(".//HDon")
        if hdon_node is not None:
            invoice_node = hdon_node.find("DLHDon")
        if invoice_node is None:
            for tag in [".//DLHDon", ".//TDiep", ".//Invoice"]:
                node = root.find(tag)
                if node is not None:
                    invoice_node = node
                    break
        if invoice_node is None:
            return None

        def get(path):
            node = invoice_node
            for part in path.split("/"):
                if node is not None:
                    node = node.find(part)
                else:
                    return None
            return node.text.strip() if node is not None and node.text else None

        bank_acc = get("NDHDon/NBan/STKNHang")
        if not bank_acc:
            for item in invoice_node.findall(".//NBan/TTKhac/TTin"):
                if item.findtext("TTruong") == "SellerBankAccount":
                    bank_acc = item.findtext("DLieu")
                    break

        return {
            "InvoiceNumber": get("TTChung/SHDon"),
            "SellerName": get("NDHDon/NBan/Ten"),
            "SellerTax": get("NDHDon/NBan/MST"),
            "SellerAddress": get("NDHDon/NBan/DChi"),
            "SellerBankAccount": bank_acc,
            "BuyerName": get("NDHDon/NMua/Ten"),
            "BuyerAddress": get("NDHDon/NMua/DChi"),
            "BuyerTax": get("NDHDon/NMua/MST")
        }

    except Exception as e:
        print(f"Parse error: {e}")
        return None


def save_to_excel(excel_path, row):
    if not os.path.isfile(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"
        ws.append([
            "No", "TaxCode", "LookupCode", "URL",
            "InvoiceNumber", "SellerName", "SellerTax",
            "SellerAddress", "SellerBankAccount",
            "BuyerName", "BuyerAddress", "BuyerTax"
        ])
        wb.save(excel_path)

    wb = load_workbook(excel_path)
    ws = wb.active
    ws.append(row)
    wb.save(excel_path)


def main():
    input_file = "input.xlsx"
    output_file = "output.xlsx"
    base_dir = os.path.join(os.getcwd(), "Invoices")

    driver, wait = setup_browser(base_dir)

    df = pd.read_excel(input_file, dtype=str)

    for idx, row in df.iterrows():
        no = idx + 1
        tax_code = str(row.get("Mã số thuế", "")).strip()
        lookup_code = str(row.get("Mã tra cứu", "")).strip()
        url = str(row.get("URL", "")).strip()

        if not url or not lookup_code:
            continue

        print(f"[{no}] Lookup {lookup_code} at {url}")

        lookup_invoice(driver, wait, tax_code, lookup_code, url)
        xml_path = download_xml(driver, wait, base_dir, url, lookup_code)

        if xml_path:
            data = parse_xml(xml_path)
            if data:
                line = [no, tax_code, lookup_code, url] + list(data.values())
            else:
                line = [no, tax_code, lookup_code, url] + [""] * 8
        else:
            line = [no, tax_code, lookup_code, url] + [""] * 8

        save_to_excel(output_file, line)

    driver.quit()
    print(f"Done. Output saved to {output_file}")


if __name__ == "__main__":
    main()
